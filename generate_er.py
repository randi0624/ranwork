import openpyxl
import json
import re

wb = openpyxl.load_workbook(r'e:\workspace\project\数据结构相关\数仓数据字典-V1.1.3-20260611.xlsx')

# 1. Read directory sheet for table comments
dir_ws = wb['数据字典目录']
table_comments = {}
table_sources = {}
for r in range(2, dir_ws.max_row + 1):
    tbl = dir_ws.cell(r, 7).value
    if tbl:
        tbl = str(tbl).strip()
        table_comments[tbl] = str(dir_ws.cell(r, 3).value or '').strip()
        table_sources[tbl] = str(dir_ws.cell(r, 5).value or '').strip()

# 2. Layer mapping
def get_layer(tname):
    t = tname.lower()
    if t.startswith('dim_'): return 'DIM 维表层'
    if t.startswith('dwd_'): return 'DWD 明细层'
    if t.startswith('dws_'): return 'DWS 汇总层'
    if t.startswith('ads_'): return 'ADS 应用层'
    # event tables and others
    for prefix in ['alarm_', 'adas_', 'risky_', 'vehicle_', 'possible_', 'park_', 'trip_', 'charing_']:
        if t.startswith(prefix): return 'ADS 应用层'
    return 'ADS 应用层'

layer_order = {'DIM 维表层': 0, 'DWD 明细层': 1, 'DWS 汇总层': 2, 'ADS 应用层': 3}

# 3. Read all table sheets
skip_sheets = ['更新记录', '数据字典目录']
sheet_names = [s for s in wb.sheetnames if s not in skip_sheets]

# D structure: D[layer][table_name] = {comment, db, columns: [...]}
D = {}
total_fields = 0

for sn in sheet_names:
    ws = wb[sn]
    
    # Extract db.table from R2
    r2_val = str(ws.cell(2, 2).value or '')
    db_name = ''
    table_name = sn  # default to sheet name
    
    m = re.search(r'库名\+表名[：:]\s*(\S+)', r2_val)
    if m:
        full_name = m.group(1)
        if '.' in full_name:
            parts = full_name.split('.', 1)
            db_name = parts[0]
            table_name = parts[1]
        else:
            table_name = full_name
    
    # Clean table name (remove trailing dots or spaces)
    table_name = table_name.strip().rstrip('.')
    
    layer = get_layer(table_name)
    
    # Parse fields from R3+
    columns = []
    for r in range(3, ws.max_row + 1):
        seq = ws.cell(r, 2).value
        fname = ws.cell(r, 3).value
        ftype = ws.cell(r, 4).value
        fcomment = ws.cell(r, 5).value
        fremark = ws.cell(r, 6).value
        
        if not fname:
            continue
        fname = str(fname).strip()
        if not fname or fname == 'None':
            continue
            
        ftype = str(ftype or '').strip()
        fcomment = str(fcomment or '').strip()
        fremark = str(fremark or '').strip()
        
        comment_full = fcomment
        if fremark and fremark != 'None':
            comment_full = f"{fcomment} | {fremark}" if fcomment else fremark
        
        columns.append({
            'name': fname,
            'type': ftype,
            'key': '',
            'nullable': 'YES',
            'default': '',
            'comment': comment_full
        })
    
    total_fields += len(columns)
    
    # Build comment from directory
    comment = table_comments.get(table_name, '')
    if db_name:
        comment = f"[{db_name}] {comment}" if comment else f"[{db_name}]"
    
    if layer not in D:
        D[layer] = {}
    
    D[layer][table_name] = {
        'comment': comment,
        'db': db_name,
        'columns': columns
    }

wb.close()

# 4. Stats
total_tables = sum(len(v) for v in D.values())
print(f"Layers: {len(D)}")
for layer in sorted(D.keys(), key=lambda x: layer_order.get(x, 99)):
    print(f"  {layer}: {len(D[layer])} tables")
print(f"Total tables: {total_tables}")
print(f"Total fields: {total_fields}")

# 5. Generate JS data
# Sort tables within each layer
for layer in D:
    D[layer] = dict(sorted(D[layer].items()))

# Build JS
layers_sorted = sorted(D.keys(), key=lambda x: layer_order.get(x, 99))

js_DL = json.dumps({l: l for l in layers_sorted}, ensure_ascii=False)
js_DBcls = json.dumps({l: ['b-dwd','b-dws','b-dwd','b-ads'][layer_order.get(l,3)] for l in layers_sorted}, ensure_ascii=False)
js_DBcolor = json.dumps({l: ['#0984e3','#00a878','#0984e3','#6c5ce7'][layer_order.get(l,3)] for l in layers_sorted}, ensure_ascii=False)

# Serialize D
js_D = json.dumps(D, ensure_ascii=False)

# 6. Generate HTML
html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>数仓 ER 视图 - V1.1.3</title>
<style>
:root{--bg:#f5f7fb;--panel:#ffffff;--ink:#1f2937;--muted:#667085;--line:#9aa6b2;--accent:#0f766e;--accent2:#2563eb;--border:#d9e0ea;--soft:#eef6f5;--warn:#b45309;--shadow:0 10px 28px rgba(31,41,55,.10)}
*{box-sizing:border-box}
body{margin:0;font-family:Segoe UI,Microsoft YaHei,Arial,sans-serif;color:var(--ink);background:var(--bg);font-size:14px}
.app{height:100vh;display:grid;grid-template-columns:280px 1fr;overflow:hidden}

/* Sidebar */
.sidebar{background:var(--panel);border-right:1px solid var(--border);overflow:auto;display:flex;flex-direction:column}
.sidebar::-webkit-scrollbar{width:5px}
.sidebar::-webkit-scrollbar-thumb{background:#ccc;border-radius:3px}
.brand{padding:18px 18px 12px;border-bottom:1px solid var(--border)}
.brand h1{font-size:18px;line-height:1.25;margin:0 0 8px}
.brand p{margin:0;color:var(--muted);line-height:1.5;font-size:13px}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:14px}
.stat{background:#f8fafc;border:1px solid var(--border);border-radius:8px;padding:8px}
.stat b{display:block;font-size:18px}
.stat span{font-size:12px;color:var(--muted)}
.side-section{padding:14px 14px 18px;flex:1}
.search{width:100%;height:36px;border:1px solid var(--border);border-radius:8px;padding:0 10px;background:#fff;color:var(--ink);font-size:13px;outline:none}
.search:focus{border-color:var(--accent)}
.domain-list{display:flex;flex-direction:column;gap:8px;margin-top:12px}
.domain-btn{border:1px solid var(--border);background:#fff;color:var(--ink);border-radius:8px;padding:10px 11px;text-align:left;cursor:pointer;font-weight:600;font-size:13px;transition:all .15s}
.domain-btn:hover{border-color:#8aa1bd}
.domain-btn.active{background:var(--soft);border-color:#6bb6ad;color:#075e57}
.domain-btn small{float:right;color:var(--muted);font-weight:500}
.legend{margin-top:18px;color:var(--muted);font-size:12px;line-height:1.6}

/* Main canvas */
.main{display:grid;grid-template-rows:auto 1fr;background:linear-gradient(#f8fafc,#f3f6fb);overflow:hidden}
.toolbar{height:62px;display:flex;align-items:center;gap:10px;padding:12px 16px;border-bottom:1px solid var(--border);background:rgba(255,255,255,.82);backdrop-filter:blur(8px);flex-shrink:0}
.toolbar h2{font-size:18px;margin:0 12px 0 0;min-width:100px}
.col-setting{margin-left:auto;display:flex;align-items:center;gap:6px;font-size:12px;color:var(--muted)}
.col-setting input{width:42px;height:28px;border:1px solid var(--border);border-radius:6px;padding:0 4px;font-size:12px;text-align:center;outline:none}
.col-setting input:focus{border-color:var(--accent)}
.canvas-wrap{position:relative;overflow:auto;min-height:0}
.canvas-wrap::-webkit-scrollbar{width:8px;height:8px}
.canvas-wrap::-webkit-scrollbar-thumb{background:#bbb;border-radius:4px}
.canvas-wrap::-webkit-scrollbar-track{background:#f0f2f5}
.canvas{padding:24px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:20px}

/* Table card */
.table-card{background:var(--panel);border:1px solid var(--border);border-radius:8px;box-shadow:var(--shadow);overflow:hidden;transition:box-shadow .15s}
.table-card:hover{box-shadow:0 12px 32px rgba(31,41,55,.14)}
.table-head{padding:10px 12px;background:#f8fafc;border-bottom:1px solid var(--border);cursor:pointer}
.table-name{font-size:13px;font-weight:700;color:#111827;word-break:break-all}
.table-comment{font-size:12px;color:var(--muted);margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.fields{list-style:none;margin:0;padding:6px 0}
.field{display:grid;grid-template-columns:22px 1fr;gap:7px;padding:4px 10px;border-bottom:1px solid #f1f4f8;min-height:28px}
.field:last-child{border-bottom:0}
.badge{width:20px;height:20px;border-radius:5px;display:inline-flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;margin-top:1px}
.plain{background:#edf2f7;color:#64748b}
.field-main{min-width:0}
.field-name{font-size:12px;font-weight:600;word-break:break-all}
.field-meta{font-size:11px;color:var(--muted);margin-top:2px;line-height:1.35}
.more{padding:7px 10px;color:var(--muted);font-size:12px;background:#fafbfc;cursor:pointer;text-align:center;border-top:1px solid #f1f4f8}
.more:hover{background:#f0f2f5}

/* Layer badges */
.b-dim{background:#fff3e8;color:#e17055}
.b-dwd{background:#e8f4fd;color:#0984e3}
.b-dws{background:#e6f9f0;color:#00a878}
.b-ads{background:#f3f0ff;color:#6c5ce7}

/* Modal */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.35);z-index:100;display:none;align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal{background:#fff;border-radius:8px;width:92vw;max-width:900px;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 12px 40px rgba(0,0,0,.18)}
.modal-hd{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.modal-hd .mh-title{font-size:16px;font-weight:700;flex:1}
.modal-hd .mh-badge{font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600}
.modal-close{background:none;border:none;font-size:22px;color:var(--muted);cursor:pointer;padding:0 4px;line-height:1}
.modal-close:hover{color:var(--ink)}
.modal-body{flex:1;overflow-y:auto;padding:0}
.modal-body::-webkit-scrollbar{width:6px}
.modal-body::-webkit-scrollbar-thumb{background:#ccc;border-radius:3px}
.modal-cmt{padding:4px 18px 8px;font-size:12px;color:var(--muted)}
.field-table{width:100%;border-collapse:collapse;font-size:12px}
.field-table th,.field-table td{border-bottom:1px solid #edf1f6;padding:7px 6px;text-align:left;vertical-align:top}
.field-table th{color:#475467;background:#f8fafc;position:sticky;top:0;z-index:1}
.field-table code{font-family:Consolas,Menlo,monospace;font-size:11px}
.pill{display:inline-block;border:1px solid var(--border);border-radius:999px;padding:2px 7px;font-size:11px;background:#fff}
</style>
</head>
<body>
<div class="app">
  <aside class="sidebar">
    <div class="brand">
      <h1>数仓 ER 视图</h1>
      <p>车辆数据平台数仓表结构 V1.1.3</p>
      <div class="stats">
        <div class="stat"><b id="sDb">0</b><span>层级</span></div>
        <div class="stat"><b id="sTbl">0</b><span>表</span></div>
        <div class="stat"><b id="sCol">0</b><span>字段</span></div>
      </div>
    </div>
    <div class="side-section">
      <input id="search" class="search" placeholder="搜索表名 / 字段 / 注释" oninput="doSearch(this.value)">
      <div id="domains" class="domain-list"></div>
      <div class="legend"><b>层级说明</b><br>DIM：维表层<br>DWD：数据明细层<br>DWS：数据汇总层<br>ADS：数据应用层<br>点击表名查看字段详情</div>
    </div>
  </aside>
  <main class="main">
    <div class="toolbar">
      <h2 id="domainTitle">总览</h2>
      <div class="col-setting">概览字段数 <input type="number" id="colNum" min="1" max="30" value="6" onchange="chgCols(this.value)"></div>
    </div>
    <div class="canvas-wrap">
      <div class="canvas"><div class="grid" id="grid"></div></div>
    </div>
  </main>
</div>

<div class="modal-overlay" id="mo" onclick="closeModal(event)">
  <div class="modal" onclick="event.stopPropagation()">
    <div class="modal-hd">
      <span class="mh-badge" id="mb"></span>
      <span class="mh-title" id="mt"></span>
      <button class="modal-close" onclick="closeModal()">&times;</button>
    </div>
    <div class="modal-cmt" id="mcmt"></div>
    <div class="modal-body" id="mbd"></div>
  </div>
</div>

<script>
const D=''' + js_D + ''';
const DL=''' + js_DL + ''';
const DBcls=''' + json.dumps({l: ['b-dim','b-dwd','b-dws','b-ads'][layer_order.get(l,3)] for l in layers_sorted}, ensure_ascii=False) + ''';
const DBcolor=''' + js_DBcolor + ''';

let showCols=6, curFilter=null, curTable=null;
const DN=Object.keys(D).sort((a,b)=>{
  const order=''' + json.dumps({l: layer_order.get(l,99) for l in layers_sorted}) + ''';
  return (order[a]||99)-(order[b]||99);
});
let tt=0,tc=0;
DN.forEach(d=>{const ts=Object.keys(D[d]);tt+=ts.length;ts.forEach(t=>tc+=D[d][t].columns.length)});
document.getElementById('sDb').textContent=DN.length;
document.getElementById('sTbl').textContent=tt;
document.getElementById('sCol').textContent=tc;

const domDiv=document.getElementById('domains');
const allBtn=document.createElement('button');
allBtn.className='domain-btn active';allBtn.innerHTML='全部 <small>'+tt+' 表</small>';
allBtn.onclick=()=>filterDomain(null,allBtn);domDiv.appendChild(allBtn);
DN.forEach(d=>{
  const cnt=Object.keys(D[d]).length;
  const btn=document.createElement('button');
  btn.className='domain-btn';
  btn.innerHTML=(DL[d]||d)+' <small>'+cnt+' 表</small>';
  btn.onclick=()=>filterDomain(d,btn);
  domDiv.appendChild(btn);
});

function filterDomain(d,btn){
  curFilter=d;
  document.querySelectorAll('.domain-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('domainTitle').textContent=d?(DL[d]||d):'总览';
  renderGrid();
}

function renderGrid(){
  const grid=document.getElementById('grid');
  let h='';
  DN.forEach(d=>{
    if(curFilter&&curFilter!==d)return;
    const ts=Object.keys(D[d]).sort();
    ts.forEach(tn=>{
      const info=D[d][tn],cols=info.columns,show=cols.slice(0,showCols),rest=cols.length-showCols;
      h+='<div class="table-card" data-db="'+d+'" data-tbl="'+tn+'">';
      h+='<div class="table-head" onclick="selectTable(&#39;'+d+'&#39;,&#39;'+tn+'&#39;)">';
      h+='<div class="table-name" style="color:'+DBcolor[d]+'">'+tn+'</div>';
      if(info.comment)h+='<div class="table-comment" title="'+esc(info.comment)+'">'+esc(info.comment)+'</div>';
      h+='</div>';
      h+='<ul class="fields">';
      show.forEach(c=>{
        h+='<li class="field"><span class="badge plain">·</span><div class="field-main"><div class="field-name">'+esc(c.name)+'</div>';
        h+='<div class="field-meta"><code>'+esc(c.type)+'</code>';
        if(c.comment)h+=' · '+esc(c.comment);
        h+='</div></div></li>';
      });
      h+='</ul>';
      if(rest>0)h+='<div class="more" onclick="selectTable(&#39;'+d+'&#39;,&#39;'+tn+'&#39;)">还有 '+rest+' 个字段 →</div>';
      h+='</div>';
    });
  });
  grid.innerHTML=h;
}

function selectTable(d,tn){
  curTable={db:d,tbl:tn};
  document.querySelectorAll('.table-card').forEach(c=>{c.style.outline='';c.style.borderColor=''});
  const card=document.querySelector('.table-card[data-db="'+d+'"][data-tbl="'+tn+'"]');
  if(card){card.style.outline='3px solid rgba(37,99,235,.22)';card.style.borderColor='#6096f5'}
  openModal(d,tn);
}

function buildFieldTable(cols){
  let h='<table class="field-table"><thead><tr><th>#</th><th>字段名</th><th>类型</th><th>注释</th></tr></thead><tbody>';
  cols.forEach((c,i)=>{
    h+='<tr><td style="color:#aaa;font-size:11px">'+(i+1)+'</td>';
    h+='<td><code>'+esc(c.name)+'</code></td>';
    h+='<td><code>'+esc(c.type)+'</code></td>';
    h+='<td>'+esc(c.comment)+'</td></tr>';
  });
  return h+'</tbody></table>';
}

function openModal(d,tn){
  const info=D[d][tn],bcls=DBcls[d]||'b-ads';
  document.getElementById('mb').className='mh-badge '+bcls;
  document.getElementById('mb').textContent=DL[d]||d;
  document.getElementById('mt').textContent=tn;
  document.getElementById('mcmt').textContent=info.comment||'';
  document.getElementById('mbd').innerHTML=buildFieldTable(info.columns);
  document.getElementById('mo').classList.add('open');
  document.body.style.overflow='hidden';
}
function closeModal(e){
  if(e&&e.target!==document.getElementById('mo'))return;
  document.getElementById('mo').classList.remove('open');
  document.body.style.overflow='';
}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeModal()});

function chgCols(v){
  const n=parseInt(v);if(n>=1&&n<=50){showCols=n;renderGrid()}
}

function doSearch(v){
  v=v.toLowerCase().trim();
  document.querySelectorAll('.table-card').forEach(card=>{
    const tn=card.getAttribute('data-tbl').toLowerCase();
    const d=card.getAttribute('data-db');
    let cm=false;
    if(D[d]&&D[d][card.getAttribute('data-tbl')])cm=D[d][card.getAttribute('data-tbl')].columns.some(c=>c.name.toLowerCase().includes(v)||(c.comment||'').toLowerCase().includes(v));
    card.style.display=(!v||tn.includes(v)||cm)?'':'none';
  });
}

function esc(s){if(!s)return '';return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}

renderGrid();
</script>
</body>
</html>'''

# Fix escaped quotes in onclick (Python string escaping issue)
html = html.replace("\\'", "'")

output_path = r'e:\workspace\project\数据结构相关\数仓ER视图.html'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\nHTML generated: {output_path}")
print(f"File size: {len(html):,} chars")
