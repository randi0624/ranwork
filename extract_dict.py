import openpyxl
import json

wb = openpyxl.load_workbook(r'e:\workspace\project\数据结构相关\数仓数据字典-V1.0-20260605.xlsx')

# Skip first 2 sheets (更新记录, 数据字典目录)
skip_sheets = ['更新记录', '数据字典目录']
sheet_names = [s for s in wb.sheetnames if s not in skip_sheets]

print(f"Total sheets to process: {len(sheet_names)}")

# Read each sheet and output header structure + first data row
for sn in sheet_names[:3]:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
    for r in range(1, min(4, ws.max_row+1)):
        row_data = []
        for c in range(1, min(ws.max_column+1, 12)):
            val = ws.cell(r, c).value
            row_data.append(str(val)[:50] if val else '')
        print(f"  R{r}: {row_data}")

wb.close()
