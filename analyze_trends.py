import openpyxl

files = [
    (r'e:\workspace\project\数据相关\1-车辆数据主要字段全生命周期表_V1-永安保险截止20226.5.22.xlsx', '2026.5.22'),
    (r'e:\workspace\project\数据相关\2-车辆数据主要字段全生命周期表_V1-永安保险截止20226.5.29.xlsx', '2026.5.29'),
    (r'e:\workspace\project\数据相关\3-车辆数据主要字段全生命周期表_V1-永安保险截止20226.6.5.xlsx', '2026.6.5'),
]

# 列定义（Row2为实际表头，Row3开始为数据）
col_map = {
    1: '序号', 2: '车辆编号', 3: '车驾号', 4: '数据授权日期', 5: '数据接入日期',
    6: '总保费', 7: '车龄', 8: '故障记录', 9: '驾驶行为评分', 10: '健康评分',
    11: '智能驾驶预警', 12: '总里程', 13: 'AEB里程', 14: 'ACC里程',
    15: '总充电次数', 16: '低电量充电', 17: '故障充电', 18: '累计停车',
    19: '累计急加速', 20: '累计急减速', 21: '累计急转弯',
    22: '累计分心驾驶', 23: '累计疲劳驾驶', 24: '累计危险驾驶',
    25: '累计FCW', 26: '累计AEB', 27: '累计LDP', 28: '备注'
}

# 数值型列
num_cols = [6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27]

all_data = {}
for fpath, label in files:
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    rows_data = []
    for r in range(3, ws.max_row + 1):
        vid = ws.cell(r, 2).value
        if vid is None:
            continue
        row = {}
        for c in range(1, 29):
            val = ws.cell(r, c).value
            if c in num_cols and val is not None:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = 0
            elif c in num_cols:
                val = 0
            row[c] = val
        rows_data.append(row)
    all_data[label] = rows_data
    wb.close()

# ===== 1. 基础统计 =====
print("=" * 80)
print("一、基础概览")
print("=" * 80)
for label in ['2026.5.22', '2026.5.29', '2026.6.5']:
    data = all_data[label]
    print(f"\n【{label}】车辆数: {len(data)}")
    premiums = [r[6] for r in data if r[6]]
    print(f"  总保费: {sum(premiums):.2f} 元, 均值: {sum(premiums)/len(premiums):.2f} 元" if premiums else "  无保费数据")

# ===== 2. 各维度周对比 =====
print("\n" + "=" * 80)
print("二、各维度周度趋势对比（均值/总量）")
print("=" * 80)

dim_names = {
    8: '故障记录', 9: '驾驶行为评分', 10: '健康评分',
    11: '智能驾驶预警', 12: '总里程(KM)', 13: 'AEB里程(KM)', 14: 'ACC里程',
    15: '总充电次数', 16: '低电量充电', 17: '故障充电',
    18: '累计停车', 19: '累计急加速', 20: '累计急减速', 21: '累计急转弯',
    22: '累计分心驾驶', 23: '累计疲劳驾驶', 24: '累计危险驾驶',
    25: '累计FCW', 26: '累计AEB', 27: '累计LDP'
}

for col, name in dim_names.items():
    print(f"\n--- {name} (列{col}) ---")
    for label in ['2026.5.22', '2026.5.29', '2026.6.5']:
        data = all_data[label]
        vals = [r[col] for r in data if r[col] is not None and r[col] != 0]
        all_vals = [r[col] for r in data]
        total = sum(all_vals)
        avg = total / len(data) if data else 0
        non_zero = len(vals)
        max_v = max(all_vals) if all_vals else 0
        min_v = min(all_vals) if all_vals else 0
        print(f"  [{label}] 总量={total:.1f}, 均值={avg:.1f}, 非零车辆={non_zero}/{len(data)}, 最大={max_v:.1f}, 最小={min_v:.1f}")

# ===== 3. 周增量分析 =====
print("\n" + "=" * 80)
print("三、各维度周增量分析（按车辆编号匹配）")
print("=" * 80)

# 建立车辆索引
vehicle_maps = {}
for label in ['2026.5.22', '2026.5.29', '2026.6.5']:
    vmap = {}
    for r in all_data[label]:
        vid = str(r[2]).strip()
        vmap[vid] = r
    vehicle_maps[label] = vmap

labels = ['2026.5.22', '2026.5.29', '2026.6.5']
# 找出所有车辆
all_vehicles = set()
for vmap in vehicle_maps.values():
    all_vehicles.update(vmap.keys())

print(f"\n三周总共出现过的车辆数: {len(all_vehicles)}")

# 每周的车辆数变化
for label in labels:
    print(f"  {label}: {len(vehicle_maps[label])} 辆车")

# 新增/减少车辆
for i in range(1, len(labels)):
    prev = set(vehicle_maps[labels[i-1]].keys())
    curr = set(vehicle_maps[labels[i]].keys())
    added = curr - prev
    removed = prev - curr
    print(f"\n  {labels[i-1]} -> {labels[i]}:")
    print(f"    新增车辆: {len(added)} 辆")
    print(f"    减少车辆: {len(removed)} 辆")
    if added:
        for v in sorted(added):
            r = vehicle_maps[labels[i]][v]
            print(f"      + {v} (保费:{r[6]})")

# 持续存在的车辆的增量
print("\n" + "=" * 80)
print("四、持续在保车辆的周增量统计")
print("=" * 80)

common_w1_w2 = set(vehicle_maps['2026.5.22'].keys()) & set(vehicle_maps['2026.5.29'].keys())
common_w2_w3 = set(vehicle_maps['2026.5.29'].keys()) & set(vehicle_maps['2026.6.5'].keys())
common_all = common_w1_w2 & set(vehicle_maps['2026.6.5'].keys())

print(f"\n三周持续在保车辆: {len(common_all)} 辆")

key_dims = [12, 13, 14, 15, 16, 17, 18, 20, 21, 25, 26, 27, 11]
key_dim_names = ['总里程', 'AEB里程', 'ACC里程', '总充电', '低电量充电', '故障充电',
                 '累计停车', '累计急减速', '累计急转弯', '累计FCW', '累计AEB', '累计LDP', '智能驾驶预警']

for col, name in zip(key_dims, key_dim_names):
    print(f"\n--- {name} 周增量 ---")
    # Week1 -> Week2
    deltas_1_2 = []
    for vid in common_w1_w2:
        v1 = vehicle_maps['2026.5.22'][vid][col] or 0
        v2 = vehicle_maps['2026.5.29'][vid][col] or 0
        deltas_1_2.append(v2 - v1)
    # Week2 -> Week3
    deltas_2_3 = []
    for vid in common_w2_w3:
        v2 = vehicle_maps['2026.5.29'][vid][col] or 0
        v3 = vehicle_maps['2026.6.5'][vid][col] or 0
        deltas_2_3.append(v3 - v2)

    print(f"  W1->W2 (5.22->5.29, {len(deltas_1_2)}辆): 总增量={sum(deltas_1_2):.1f}, 均增量={sum(deltas_1_2)/len(deltas_1_2):.1f}, 正增长={sum(1 for d in deltas_1_2 if d>0)}辆")
    print(f"  W2->W3 (5.29->6.5,  {len(deltas_2_3)}辆): 总增量={sum(deltas_2_3):.1f}, 均增量={sum(deltas_2_3)/len(deltas_2_3):.1f}, 正增长={sum(1 for d in deltas_2_3 if d>0)}辆")

# ===== 5. 风险维度重点分析 =====
print("\n" + "=" * 80)
print("五、风险维度重点分析")
print("=" * 80)

# 5a. 驾驶行为评分变化
print("\n--- 驾驶行为评分趋势 ---")
for label in labels:
    scores = [vehicle_maps[label][v][9] for v in vehicle_maps[label] if vehicle_maps[label][v][9]]
    if scores:
        print(f"  [{label}] 均分={sum(scores)/len(scores):.1f}, ≥90分={sum(1 for s in scores if s>=90)}, 80-89={sum(1 for s in scores if 80<=s<90)}, 70-79={sum(1 for s in scores if 70<=s<80)}, <70={sum(1 for s in scores if s<70)}")

# 5b. 健康评分变化
print("\n--- 健康评分趋势 ---")
for label in labels:
    scores = [vehicle_maps[label][v][10] for v in vehicle_maps[label] if vehicle_maps[label][v][10]]
    if scores:
        print(f"  [{label}] 均分={sum(scores)/len(scores):.1f}, ≥90分={sum(1 for s in scores if s>=90)}, 80-89={sum(1 for s in scores if 80<=s<90)}, 70-79={sum(1 for s in scores if 70<=s<80)}, <70={sum(1 for s in scores if s<70)}")

# 5c. 高风险车辆识别
print("\n--- 高风险车辆（驾驶评分<70 或 健康评分<70）---")
for label in labels:
    risky = []
    for vid, r in vehicle_maps[label].items():
        driving = r[9] or 100
        health = r[10] or 100
        if driving < 70 or health < 70:
            risky.append((vid, driving, health, r[6]))
    print(f"  [{label}] 高风险车辆: {len(risky)} 辆")
    for vid, d, h, p in sorted(risky, key=lambda x: x[1]+x[2]):
        print(f"    {vid}: 驾驶={d}, 健康={h}, 保费={p}")

# 5d. 预警高发车辆
print("\n--- 智能驾驶预警TOP5车辆 ---")
for label in labels:
    sorted_v = sorted(vehicle_maps[label].items(), key=lambda x: x[1][11] or 0, reverse=True)[:5]
    print(f"  [{label}]")
    for vid, r in sorted_v:
        print(f"    {vid}: 预警={r[11]}, FCW={r[25]}, AEB={r[26]}, 急转弯={r[21]}")

# 5e. 低电量充电占比
print("\n--- 低电量充电占比趋势 ---")
for label in labels:
    total_charge = sum(r[15] for r in all_data[label] if r[15])
    low_charge = sum(r[16] for r in all_data[label] if r[16])
    fault_charge = sum(r[17] for r in all_data[label] if r[17])
    ratio = low_charge/total_charge*100 if total_charge else 0
    print(f"  [{label}] 总充电={total_charge:.0f}, 低电量充电={low_charge:.0f} ({ratio:.1f}%), 故障充电={fault_charge:.0f}")

print("\n" + "=" * 80)
print("分析完成")
print("=" * 80)
