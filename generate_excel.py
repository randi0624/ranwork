import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "车辆数据"

# ========== 表头（26列，根据高清图片辨识） ==========
headers = [
    "序号",               # A
    "车辆编号",           # B
    "数据授权日期",       # C
    "数据接入日期",       # D
    "总保费（元）",       # E
    "故障记录",           # F
    "最新车辆销量",       # G
    "最新车辆销售车龄",   # H
    "智能驾驶辅助总里程", # I
    "AEB里程",           # J
    "ACC里程",           # K
    "总充电次数",         # L
    "低电量充电",         # M
    "充电桩故障",         # N
    "充电次数累计",       # O
    "停车次数累计",       # P
    "急加速累计",         # Q
    "急减速累计",         # R
    "急转弯累计",         # S
    "超速累计",           # T
    "分心驾驶累计",       # U
    "疲劳驾驶累计",       # V
    "危险驾驶累计",       # W
    "FCW累计",           # X
    "AEB累计",           # Y
    "LDP激活次数",       # Z
]

# ========== 数据行（根据高清图片逐行提取） ==========
data = [
    # Row 2 (序号1) - 数据最丰富的一行
    [1, "00000009M71XMora6", "2026/3/9", "2026/3/9",
     12796.43, 0, 83, "87*3.0", 9145, 19697.5, 19697.5,
     2.9, 171, 15, 0, 1518, 0, 6, 590, 0, 0, 0, 0, 163, 1, 3],
    # Row 3 (序号2)
    [2, "00000060S10ZWxXu", "2026/3/11", "2026/3/11",
     12742.81, 0, 85, "96*3.0", 11295, 28403.1, 28403.1,
     333.2, 206, 24, 0, 1559, 0, 23, 790, 0, 0, 0, 0, 137, 0, 0],
    # Row 4 (序号3)
    [3, "000000E0N42YMor4", "2026/3/11", "2026/3/11",
     12752.03, 0, 97, "91*3.0", 14533, 28198, 28198,
     14.5, 70, 32, 0, 936, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 5 (序号4)
    [4, "00000020K71XMorG", "2026/5/12", "2026/5/12",
     12938.8, 0, 37, "135*3.0", 9632, 20347.4, 20347.4,
     4.5, 40, 10, 0, 831, 0, 0, 297, 0, 0, 0, 0, 48, 0, 0],
    # Row 6 (序号5)
    [5, "00000060T10ZMor0", "2026/3/9", "2026/3/9",
     5251.16, 0, 153, "168*3.0", 2637, 3764.9, 3764.9,
     0.6, 8, 4, 0, 202, 0, 0, 68, 0, 0, 0, 0, 0, 0, 0],
    # Row 7 (序号6)
    [6, "00000080B71XMorQ", "2026/3/11", "2026/3/11",
     5251.16, 0, 525, "175*3.0", 5828, 6944.7, 6944.7,
     0.3, 4, 2, 0, 311, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 8 (序号7)
    [7, "00000040A71XMorA", "2026/3/11", "2026/3/11",
     5251.16, 0, 484, "181*3.0", 0, 0, 0,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 9 (序号8)
    [8, "00000020L71XMora", "2026/3/9", "2026/3/9",
     5251.16, 0, 16, "129*3.0", 1050, 1240.5, 1240.5,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 10 (序号9)
    [9, "00000040C71XMorB", "2026/3/11", "2026/3/11",
     5251.16, 0, 37, "127*3.0", 0, 0, 0,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 11 (序号10)
    [10, "00000080D71XMorC", "2026/3/11", "2026/3/11",
     5251.16, 0, 1, "96*1.0", 0, 0, 0,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 12 (序号11)
    [11, "00000060U10ZMor1", "2026/3/11", "2026/3/11",
     5251.16, 0, 5, "93*1.0", 0, 0, 0,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 13 (序号12)
    [12, "00000040E71XMorD", "2026/3/9", "2026/3/9",
     5251.16, 0, 15, "91*1.0", 0, 0, 0,
     0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    # Row 14 (序号13)
    [13, "000000AWOc5ObR4ia", "2026/6/3", "2026/6/3",
     6223.86, 0, 0, "0/0", 0, 0, 0,
     0, 0, 8, 1, 0, 93, 0, 4, 70, 0, 0, 0, 3, 0, 0],
]

# 写入表头
header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 写入数据
data_align = Alignment(horizontal="center", vertical="center")
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = data_align
        cell.border = thin_border

# 自动调整列宽
for col_idx in range(1, len(headers) + 1):
    max_len = len(str(headers[col_idx - 1]))
    for row_idx in range(2, len(data) + 2):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

# 冻结首行
ws.freeze_panes = "A2"

output_path = r"e:\workspace\project\车辆数据表.xlsx"
wb.save(output_path)
print(f"Excel 文件已保存至: {output_path}")
print(f"共 {len(headers)} 列, {len(data)} 行数据")
